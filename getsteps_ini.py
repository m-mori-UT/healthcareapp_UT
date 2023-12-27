# %%
import sys, os, time, csv, re, shutil
import linecache
import math
import scipy.stats as stats
import glob
import configparser
import argparse
import traceback
import openpyxl
import itertools
import asyncio

import numpy as np
import cv2
import pyocr, pyocr.builders
import matplotlib.pyplot as plt
from PIL import Image
plt.gray()
# %%
#OCRで使うソフトウェアを設定
#tesseract win binary(64bit)を導入
#pyocrに対応したOCRソフトとしてtesseractだけがインストールされていることを想定する
#pyocrにTesseractを指定する
pyocr.tesseract.TESSERACT_CMD = r'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'
TOOLS = pyocr.get_available_tools()
TOOL = TOOLS[0]

# %%

#コマンドライン引数
parser = argparse.ArgumentParser()
parser.add_argument('config_file', help='設定ファイルのパス')
args = parser.parse_args()

#設定読取
config = configparser.ConfigParser()
config.read(args.config_file, encoding='utf-8')
_DATA_DIR_PATH = config['PATH']['Data_dir']
_RESULT_CSV_PATH = config['PATH']['Result_csv_file']
_ERROR_CSV_PATH = config['PATH']['Error_csv_file']
_OUTPUT_EXCEL_PATH = config['PATH']['Output_excel_file']
_IMG_PERIOD_DIR = config['PATH']['Img_period_dir']
_IMG_LABEL_DIR = config['PATH']['Img_label_dir']
_IMG_SUCCESS_DIR = config['PATH']['Img_success_dir']
_IMG_ERROR_DIR = config['PATH']['Img_error_dir']
_IMG_OLDLAYOUT_DIR = config['PATH']['Img_oldlayout_dir']
# %%

class StepsImage():#Attributes:→クラス内で定義された変数
    """歩数の画像クラス

    Attributes:
        filepath (str): 画像ファイルのパス
        img_org (numpy.array): OpenCVで読み込んだ画像ファイルのデータ(カラー)
        img_org_gray (numpy.array): img_orgをグレースケール変換した
        is_dark_mode (boolean): 画像ファイルがダークモードか?
        top_region (list): 画像ファイルのグラフより上の部分(平均歩数や期間が書いているところ)
        のy座標From, To
        graph_region (list): 画像ファイルのグラフ部分のy座標From, To
        img_top (numpy.array): グラフより上の部分の画像(カラー)
        img_graph (numpy.array): グラフ部分の画像(カラー)
        img_bin_top (numpy.array): グラフより上の部分の画像(白黒２値)
        img_bin_graph (numpy.array): グラフ部分の画像(白黒２値)
    """

    BIN_BLACK = 0
    BIN_WHITE = 255
    #グラフにつき31個のbinがある
    BIN_NUMBER = 31

    def __init__(self, filepath):
        """コンストラクタ

        Arguments:
            filepath (str): ファイルパス
        """

        #何度も呼び出し、さほどデータサイズも大きくないデータはクラス変数とする
        self.filepath = filepath
        #self.filename = os.path.basename(self.filepath)
        self.fileext = os.path.splitext(self.filepath)[1].lower()
        #cv2.imreadは日本語ファイル名を読めない
        #self.img_org = cv2.imread(self.filepath)
        self.img_org = cv2.imdecode(np.fromfile(self.filepath, dtype=np.uint8), cv2.IMREAD_COLOR)
        # img_hsv = cv2. cvtColor(img_org, cv2. COLOR_BGR2HSV) #後に回す：明るさだけ変える方法を調べる
        # self.img_org = cv2.cvtColor(img_hsv-img_hsv.min())/
        # (img_hsv.max()-img_hsv.min(), cv2.COLOR_HSV2BGR)
        #読み込んだ画像をグレースケール形式に変換する
        self.img_org_gray = cv2.cvtColor(self.img_org, cv2.COLOR_BGR2GRAY)
        #最初のイメージの背景での上がどこにあるか検索している
        self.is_old_layout = self.__is_old_layout()
        self.is_dark_mode = self.__is_dark_mode()
        #最初のイメージの背景での下がどこにあるか検索している
        self.top_region, self.graph_region = self.__get_region_y_range()
        self.img_top = self.img_org[self.top_region[0]:self.top_region[1], :]
        self.img_graph = self.img_org[self.graph_region[0]:self.graph_region[1], :]
        #OCRの精度を上げるためグラフ解析用とOCR用のイメージは別の条件で2値化イメージを作る
        self.img_bin_top, self.img_bin_top_ocr = self.__get_binary_image(
            self.img_org_gray[self.top_region[0]:self.top_region[1], :]
        )
        self.img_bin_graph, self.img_bin_graph_ocr = self.__get_binary_image(
            self.img_org_gray[self.graph_region[0]:self.graph_region[1], :],
        )
        self.bar_x_ranges, self.graph_x_range, self.graph_y_range = self.__graph_info()

    def __is_old_layout(self):
        """旧レイアウト（iOS 12以前）か？

        Returns
            bool: True 旧レイアウト, False 新レイアウト
        """

        # ピンク系の色が存在している場合、旧レイアウトとみなす
        # BGRーHSV色空間の変換：色相(H),彩度(S),明度(V)
        img_hsv = cv2.cvtColor(self.img_org, cv2.COLOR_BGR2HSV)
        img_pink = np.where((img_hsv[:,:,0] > 120) & (img_hsv[:,:,1] > 160) & (img_hsv[:,:,2] > 220) , 0, 255)
        cnt = np.count_nonzero(img_pink == 0)
        if cnt > 500:
            return True
        return False

    def __is_dark_mode(self):
        """画像はダークモードか?

        Returns
            bool: True ダークモード, False 通常モード
        """
        # ↓これは前のやつのとき白がダークと思われたかも？色みを替えたコード
        #_, img = cv2.threshold(self.img_org_gray, 254, 255, cv2.THRESH_BINARY)
        # 閾値以下の値を「0」それ以外の値を「最大値」として変換する方法？
        _, img = cv2.threshold(self.img_org_gray, 128, 255, cv2.THRESH_BINARY)
        return (cv2.countNonZero(img)/img.size < 0.3)

    def __get_binary_image(self, img):
        """白黒2値データを取得する

        Arguments:
            img (numpy.array): 画像データ(カラー)

        Returns:
            numpy.array -- 白黒2値データ
            numpy.array -- OCR用の白黒2値データ
        """
        # 250のthresholdを上げると、ノイズのピクセルが多くなる
        # jpegでは線と数字の周りに背景のはずなピクセルが少し暗くなるので
        # thresholdに入り、ノイズピクセルになってしまう
        # ですが、thresholdを下げたら、読取るべきな薄い灰色な線が読取れなくなってしまう
        th_min, th_max, bin_mode = (250, 255, cv2.THRESH_BINARY) if not self.is_dark_mode else (2, 255, cv2.THRESH_BINARY_INV)
        th_min_ocr, th_max_ocr, bin_mode_ocr = (230, 255, cv2.THRESH_BINARY) if not self.is_dark_mode else (35, 255, cv2.THRESH_BINARY_INV)

        _, img_bin = cv2.threshold(img, th_min, th_max, bin_mode)
        _, img_bin_ocr = cv2.threshold(img, th_min_ocr, th_max_ocr, bin_mode_ocr)

        return img_bin, img_bin_ocr

    def __get_region_y_range(self):
        """画像中のグラフ部分とその上のヘッダー部分を引き分けて、引き分けの高さ（ピクセル）をリターンする。

        Returns:
            (list, list) -- ([グラフの一番上の横線y座標から、写真の一番上まで], [グラフ下の横線から、グラフの一番上の横線のy座標まで])
        """

        thin_horizontal_lines = self._get_thin_horizontal_lines(self.img_org_gray)
        #一番上にある細く長い横線をtop_yに選択する
        # この方法はOS(現在OS17)にきく:グラフの一番上の横線はページ中の一番上の細く長い横線と同じ
        # __get_binary_imageのthresholdはあまい場合は、綺麗に横線を読み取れなくなる。
        #  こうなると、thin_horizontal_linesにある最初の結果はグラフの一番上の横線に被らないおそれがあって、
        #  後でエラー発生するか間違っている結果が出る可能性は高い。
        top_y = thin_horizontal_lines[0][1] if thin_horizontal_lines else 0

        if top_y < 21:
            raise ValueError('写真には横線が見つかりませんでした')

        # 棒グラフ下のy座標を見つける方法:一番大きいオレンジの範囲のすぐ下の座標を使う
        #  一番大きい範囲を選択する理由は画像内にグラフ以外のオレンジ色がある可能性に対応するため

        # オレンジの範囲を取得する
        orange_rep_points = self._get_orange_regions(self.img_org)
        # 一番高さがあるオレンジ範囲（棒グラフの範囲）を選択し、
        #  その棒グラフの最下部のy座標を取得する
        _, bottom_y = max(orange_rep_points, key=lambda r: r[1] - r[0])

        # グラフの一番上の横線とグラフの下の線を使って、画像を引き分けてリターンする。
        # 右にある最大数値がグラフの範囲に入るように、引き分け位置は実際線から少しだけ上にずらす
        #   この20の数字はただいくつの写真(pngとjpeg)を見て選択した
        return [0, top_y - 1], [top_y + 1 - 20, bottom_y + 10 - 1]

    def _get_thin_horizontal_lines(self, img_gray):
        """写真から縦幅が細く、横に長い線の位置（即ちピクセル高さ;y座標)をりーたんする
        """

        #デバックしやすいようにイメージをselfに保存
        self.debug_image_lines, _ = self.__get_binary_image(img_gray)

        #横長い線（のy座標）を取得
        #横が長い行の認識方法: 一行にある黒いピクセルは50%を超えたら→横長いと認識
        #  注意: この0.5の数字は写真によって調整する必要性が出てくるかもしれない（その場合0.5だと低い可能性がある）
        #  例えば、歩数がかなり多く棒グラフが上の線に触れるほど伸びてたら、0.5のthresholdを超える可能性がある。この場合には、数値を上げるだけ
        horizontal_line_y_coordinates = _get_lines_darker_than(self.debug_image_lines,
                                                               darkness_ratio_threshold=0.5,
                                                               colors_per_row=self.img_org.shape[1])
        
        #連続な横線をグループ化([1, 2, 115, 117, 200]→[[1, 2], [115, 117], [200]]
        rep_points = _get_repeated_points(horizontal_line_y_coordinates)
        #なんとなく細い感じがする横線グループを残す
        thin_horizontal_lines = list(filter(lambda x: x[1] - x[0] < 5, rep_points))

        return thin_horizontal_lines

    def _get_orange_regions(self, img_org):
        #オレンジの色を判別している
        self.img_orange, bin_black = orange_other_binarization(img_org)
        #オレンジ色がある範囲をグループ化している
        orange_rep_points = _get_repeated_points(bin_black[0])
        return orange_rep_points

    def __graph_info(self):
        """棒グラフ部分の情報を取得する

        Returns:
            (
            list,: [[棒のx軸From, To, 棒の高さ],[...],...]
            list,: [グラフ領域のx軸From, To]
            list,: [グラフ領域のy軸From, To]
            )
        """

        #バイナリ化:オレンジ→黒、他→しろ
        binary_img_graph, binary_black = orange_other_binarization(self.img_graph)

        #グラフの下の行のy座標を取得
        bottom_y = _get_lowest_graph_row(binary_black)

        graph_x_range = self._get_graph_x_range()
        bar_x_ranges = self._get_bar_x_ranges(binary_img_graph, bottom_y, binary_black)

        top_y = self._get_graph_top_y(bar_x_ranges)
        graph_y_range = [top_y, bottom_y]

        return bar_x_ranges, graph_x_range, graph_y_range

    def _get_graph_x_range(self):
        """グラフ領域のx軸範囲
        """
        # スクロールバーを無視するために、右の5%を切り落とす
        right_edge = (self.img_bin_graph.shape[1] * 95) // 100
        dark_pixel = get_black_pixels_per_line(self.img_bin_graph.transpose()[:right_edge])

        # フォントサイズが大きい場合、グラフが短くなり、数字の黒味が30%までなる可能性がある
        # それと逆、縦の線は点線の場合、60％にも行けない可能性もある。
        darkness_threshold = 0.40
        dark_lines = [r_index for r_index, pixel_count in enumerate(dark_pixel) if pixel_count / self.img_bin_graph.shape[0] >= darkness_threshold]

        #縦の線が黒ピクセル25％以上だったら線と認識している
        graph_x_range = [min(dark_lines), max(dark_lines)]
        #グラフの幅：一番左縦のｘ座標から一番右縦のｘ座標までを探している
        return graph_x_range

    def _get_bar_x_ranges(self, binary_img_graph, bottom_y, bin_black):
        """棒の高さを取得する
        """
        def get_bar_height(begin_idx, end_idx):
            center = begin_idx + int((end_idx - begin_idx) / 2)
            return bottom_y - min(bin_black[0][bin_black[1]==center]) + 1

        # bottom_y行のオレンジピクセルのｘ座標
        bottom = np.where(binary_img_graph[bottom_y, :] == self.BIN_BLACK)[0]
        #解説:左から右をループして、ちょうど棒を通ったら、その棒の高さを計算して、bar_x_rangesに入れてく
        #   具体的に、begin_idは棒の最も左な位置を示して、次の棒に飛ばす時（連続さが切られる）
        #   pre_idxは棒の最も右な位置を示している。高さの計算はbegin_idxとpre_idxの間をとって、
        #   その棒の最も高い点とbottom_yを比べてbar_x_rangesに入れる
        begin_idx, pre_idx, end_idx = bottom[0], bottom[-1], 0
        bar_x_ranges = []
        for idx in bottom:
            if (idx - pre_idx) <= 1:
                pre_idx = idx
                continue
            end_idx = pre_idx
            bar_x_ranges.append([begin_idx, end_idx, get_bar_height(begin_idx, end_idx)])
            begin_idx, pre_idx = idx, idx
        else:
            bar_x_ranges.append([begin_idx, idx, get_bar_height(begin_idx, idx)])

        return bar_x_ranges

    def _get_graph_top_y(self, bar_x_ranges):
        """グラフ領域のy軸範囲
        """
        def get_black_pixels_after_bar(bar_num):
            """
            Return: 指定された棒と右隣の棒の間にある黒ピクセル
            """
            # 最初の棒とその次の棒と間
            non_bar_center = bar_x_ranges[bar_num][1] + int((bar_x_ranges[bar_num + 1][0] - bar_x_ranges[bar_num][1])/2)
            # 棒の間にある黒のピクセル(グラフの横線のみのはず)
            return np.where(self.img_bin_graph[:, non_bar_center] == self.BIN_BLACK)[0]

        center_space_black = get_black_pixels_after_bar(bar_num=0)
        #点線がない場合は, 黒ドットは多くて30個程度だが, 点線がある場合は黒ドットは200個近く現れる
        #かなり少なめに見積もって100個を判定基準とする
        #通称:点線があった場合は多分黒ピクセルの量は100より多い
        #棒と棒の中心に縦点線が当たる場合、一つ右隣の棒と棒の中心を使う
        if 100 < len(center_space_black):
            center_space_black = get_black_pixels_after_bar(bar_num=1)

        top_y = min(center_space_black)
        return top_y


    def __repr__(self):
        info = [
            f'filepath : {self.filepath}',
        ]
        return str(info)


def _get_lowest_graph_row(binary_black):
    # グラフの一番下にある黒い色を探してるが、
    #  棒が短すぎると、オレンジ色が下の線とぶつけ合う可能性があって、
    #  max(binary_black[0])はバグってしまう (下しすぎになってしまう)
    # これを対応するため、なるべく黒が多い行の中からmaxを取る。
    #  こうすると、実際棒がある行からmaxを選択する
    # 行ごとにオレンジ色なピクセルカウント
    frequent_black = sorted(zip(*np.unique(binary_black[0], return_counts=True)), key=lambda c: (c[1], c[0]),
                            reverse=True)
    # オレンジの量上位10%多い行の中での一番下の行
    bottom_y, _ = max(frequent_black[:len(frequent_black) // 10], key=lambda x: x[0])
    return bottom_y


def _get_lines_darker_than(image, darkness_ratio_threshold, colors_per_row):
    """
    Return: 黒がdarkness_ratio_thresholdより多い行のy座標
    """
    black_pixels_per_line = get_black_pixels_per_line(image)
    return [r_index for r_index, pixel_count in enumerate(black_pixels_per_line) if darkness_ratio_threshold <= pixel_count / colors_per_row]


def _get_repeated_points(point_):
    """point_リストの中から連続の数字をグループ化している。
    連続さは30ピクセルごとに分けている

    例）[1,2,3,15,16,61,223,224,225,226,1111,1112]→[[1,16],[61],[223,226],[1111,1112]]にしている
    """
    begin_idx, pre_idx, end_idx = point_[0], point_[-1], 0
    rep_points = []
    for idx in point_:
        if (idx - pre_idx) <= 30:
            pre_idx = idx
            continue
        end_idx = pre_idx
        rep_points.append([begin_idx, end_idx])
        begin_idx, pre_idx = idx, idx
    else:
        rep_points.append([begin_idx, idx])
    return rep_points


def get_period_y_region(img_top):
    """期間のy軸From, Toを取得する

    Arguments:
        img_top (numpy.array): グラフより上の部分のデータ(白黒2値)

    Returns:
        (int, int) -- (y軸From, To)
    """

    #画像中の黒ドット部分の座標
    search_img = img_top[:, 0:int(img_top.shape[1]*2/3)]
    black = np.unique(np.where(search_img == StepsImage.BIN_BLACK)[0])

    #画像最下部から上方向に走査する
    #y軸方向に走査して初めて検出した黒ドットが続く部分が該当箇所
    pre_idx = black[black.shape[0]-1]
    for idx in black[::-1]:
        if 1 < pre_idx - idx:
            return idx - 1, img_top.shape[0] - 1
        pre_idx = idx


def get_label_y_region(img_graph, start_x, end_x):
    """棒グラフの一番上の補助線のラベルのy軸From, Toを取得する

    Arguments:
        img_graph (numpy.array): グラフ部分のデータ(白黒2値)
        start_x (int): 走査を開始するx座標

    Returns:
        (int, int): (y軸From, To)
    """
    #走査する部分を切り出した画像中の黒ドット部分の座標
    search_img = img_graph[:, start_x+2:end_x]
    black = np.unique(np.where(search_img == StepsImage.BIN_BLACK)[0])

    #画像最上部から下方向に走査する
    pre_idx = black[0]
    for idx in black:
        if 4 < idx - pre_idx:
            #多少バッファを見る
            return 0, pre_idx + 7
        pre_idx = idx
    return 0, pre_idx

def orange_other_binarization(img):
    """画像中のオレンジ系の箇所を黒, その他を白とする白黒2値に変換する

    Arguments:
        img (numpy.array): グラフ部分の画像データ(カラー)

    Returns:
        numpy.array: 白黒2値画像データ
    """

    img_hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    #オレンジ系の色（Hが0～60度）、これらの数値は結果がうまくいくようにチューニングする
    binary_img = np.where((img_hsv[:,:,0] < 20) & (img_hsv[:,:,1] > 100) & (img_hsv[:,:,2] > 100), 0, 255)
    binary_black = np.where(binary_img == StepsImage.BIN_BLACK)
    return binary_img, binary_black


def get_black_pixels_per_line(image):
    """画像内に黒ピクセルが何個あるかを行ずつで数えている"""
    color_counts_per_line = [dict(zip(*np.unique(line, return_counts=True))) for line in image]
    return [counts.get(StepsImage.BIN_BLACK, 0) for counts in color_counts_per_line]

def write_csv(path, header, data, encoding='utf8'):
    """CSV出力する

    Args:
        path (str): CSVファイルパス
        header (numpy.array): ヘッダー
        data (numpy.array): レコード
        encoding (str, optional): エンコーディング. Defaults to 'utf8'.
    """
    try:
        if os.path.exists(path):
            os.remove(path)

        with open(path, 'a', encoding=encoding, newline='') as f:
            w = csv.DictWriter(f, fieldnames=header)
            w.writeheader()
            for d in data:
                w.writerow(d)

    except IOError:
        print(f'[error]{path} is opened.')


def initialize_dir(path):
    """出力先フォルダを空にする

    Args:
        path (str): 出力先フォルダパス
    """

    try:
        if os.path.exists(path):
            shutil.rmtree(path)
        os.mkdir(path)

    except IOError:
        print(f'[error]{path} is opened.')

def extract_details(image: StepsImage):
    """メイン処理

    Arguments:
        image (StepsImage): StepsImageのインスタンス

    Returns:
        (str,:期間
         str,: 棒グラフの一番上のy軸ラベル
         list,: 棒の高さ
         int,: pixel当たりの棒の高さ(歩数/pixel)
        )
    """

    #グラフの情報を取る
    #bar_info, graph_x_range, graph_y_range = image.graph_info()
    #image.bar_x_ranges, image.graph_x_range, image.graph_y_range = image.graph_info()

    #大津の２値化
    def binarize_otsu(img):
        g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        _, o = cv2.threshold(g, 0, 255, cv2.THRESH_OTSU)
        if image.is_dark_mode:
            o = cv2.bitwise_not(o)
        return o

    period_img, period_text = _get_period_img_and_text(image, binarize_otsu)
    label_img, label_text = _get_label_img_and_text(image, binarize_otsu)
    heights = _spread_bar_heights_across_date_range(image)
    max_height_pixel = image.graph_y_range[1] - image.graph_y_range[0]

    return period_img, period_text, label_img, label_text, max_height_pixel, heights

def _get_period_img_and_text(image, binarize_otsu):

    #期間をOCRする
    period_top, period_bottom = get_period_y_region(image.img_bin_top_ocr)
    period_img = binarize_otsu(image.img_top[period_top:period_bottom, 0:int(image.img_bin_top_ocr.shape[1]*2/3)])
    period_text = TOOL.image_to_string(Image.fromarray(period_img),
                                       lang='script/Japanese',#言語は日本語
                                       builder=pyocr.builders.TextBuilder(tesseract_layout=6))#画像の文字を抽出
    #tesseract_layoutは読み込みの精度を調節するプロパティで、0から10までの値を設定できる。文字一つ一つをブロックとして認識する6に設定。
    period_text = period_text.replace(' ','').replace('~', '～')

    return period_img, period_text


def _get_label_img_and_text(image, binarize_otsu):
    #グラフの上
    # 限値をOCRする
    # グラフ右の線（点線）から写真の右いっぱいではなくその左の半分くらいまでのregion area（閾値）の中で、ラベルの上下を探す
    label_top, label_bottom = get_label_y_region(image.img_bin_graph_ocr, image.graph_x_range[1], (image.graph_x_range[1] + image.img_bin_graph_ocr.shape[1]) // 2)
    label_img = binarize_otsu(image.img_graph[label_top:label_bottom, image.graph_x_range[1]+1 :])
    label_text = TOOL.image_to_string(Image.fromarray(label_img),
                                      lang='eng',
                                      builder=pyocr.builders.DigitBuilder(tesseract_layout=7))
    label_text = label_text.replace(' ','').replace(',', '').replace('.', '')
    label_text = re.sub('[^0-9]', '', label_text)

    return label_img, label_text


def _spread_bar_heights_across_date_range(image):

    #heightが31個になるように棒のない箇所の高さをゼロとして補完する
    bar_width = image.bar_x_ranges[0][1] - image.bar_x_ranges[0][0] + 1
    graph_width = image.graph_x_range[1] - image.graph_x_range[0] - 1
    space_width = graph_width- bar_width*StepsImage.BIN_NUMBER
    space_unit = int(space_width/StepsImage.BIN_NUMBER)
    edge_space = int(space_width/StepsImage.BIN_NUMBER/2)
    b_s_width = bar_width + space_unit + 1
    j, heights = 0, np.zeros(StepsImage.BIN_NUMBER)
    for idx, bar in enumerate(image.bar_x_ranges):
        if idx == 0:
            s = bar[0] - image.graph_x_range[0] - 1
            n = int((s - edge_space)/b_s_width)
        else:
            s = bar[0] - image.bar_x_ranges[idx-1][1] - 1
            n = int((s - space_unit)/b_s_width)
        j += n
        if (j >= len(heights)): break
        heights[j] = bar[2]
        j += 1

    return heights

def extract_image_details(index, file_, files, write_xl_args, error_results, error_header):
    filename = os.path.basename(file_)
    image = None
    try:
        if not os.path.exists(file_):
            raise FileNotFoundError(file_)

        print(f'---{index+1}/{len(files)}---')
        print(filename)

        image = StepsImage(file_)

        if image.is_old_layout:
            raise AttributeError('Old layout.')
        img_period, period, img_label, label, max_height_pixel, heights = extract_details(image)
        #OCRした部分の画像を出力
        img_period_path = _IMG_PERIOD_DIR + '/' + filename
        img_label_path = _IMG_LABEL_DIR + '/' + filename
        cv2.imwrite(img_period_path, img_period)
        cv2.imwrite(img_label_path, img_label)

        write_xl_args[index] = [img_period_path, img_period, img_label_path, img_label,
                                filename, period, label, max_height_pixel, heights, file_, index]

    except FileNotFoundError as fnfe:
        error_results[index] = dict(zip(error_header, [filename, 'FileNotFoundError', '']))
        shutil.copy(file_, _IMG_ERROR_DIR)

    except Exception as e:
        _, _, tb=sys.exc_info()
        lineno=""
        while tb is not None:
            lineno +=str(tb.tb_lineno)+","
            tb=tb.tb_next
        error_results[index] = dict(zip(error_header, [filename, type(e), e.args[0] if e.args else " ", lineno]))
        if image is None:
            shutil.copy(file_, _IMG_ERROR_DIR)
            return

        if image.is_old_layout:
            shutil.copy(file_, _IMG_OLDLAYOUT_DIR)
            return
        shutil.copy(file_, _IMG_ERROR_DIR)

def write_xl_row(img_period_path, img_period, img_label_path, img_label,
                 filename, period, label, max_height_pixel, heights, file_, index,
                 ni, results_map, header, max_width_period, max_width_label, ws, wb):
    print(f'---Writing {file_}---')

    #Excelに出力
    img_period_r = openpyxl.drawing.image.Image(img_period_path)
    img_label_r = openpyxl.drawing.image.Image(img_label_path)

    #出力先のセルアドレスを取得
    now_row = ni + 1 #header行を飛ばす
    addr_period = ws.cell(row=now_row, column=1).coordinate
    addr_label = ws.cell(row=now_row, column=2).coordinate

    #セル幅調整
    h = max(img_period.shape[0], img_label.shape[0])
    ws.row_dimensions[now_row].height = h*0.75
    if max_width_period < img_period.shape[1]:
        max_width_period = img_period.shape[1]
    if max_width_label < img_label.shape[1]:
        max_width_label = img_label.shape[1]

    #画像のアンカ設定
    img_period_r.anchor = addr_period
    img_label_r.anchor = addr_label

    #画像出力
    ws.add_image(img_period_r)
    ws.add_image(img_label_r)

    #棒グラフの高さを推測
    heights_xlsf = [f'=E{now_row}/F{now_row}*{i[1]}' for i in enumerate(heights)]
    #record =  [image.filename, period, label, max_height_pixel] + heights.tolist()
    record =  [filename, period, label, max_height_pixel] + heights_xlsf

    #文字データのExcelへの埋め込み
    for i, h in enumerate(record):
        c = ws.cell(row=now_row, column=3+i)
        c.value = h

    #CSV出力用データ
    results_map.append(dict(zip(header, record)))
    shutil.copy(file_, _IMG_SUCCESS_DIR)

    #Excel Save 100件に1回
    if (index + 1) % 100 == 0:
        wb.save(_OUTPUT_EXCEL_PATH)

    return max_width_label, max_width_period

def main():
    start_time = time.time()

    #処理対象ファイルリスト
    types = ('/*.png', '/*.jpg', '/*.jpeg')
    files = []
    for type_ in types:
        files.extend(glob.glob(_DATA_DIR_PATH + type_))

    #出力先フォルダの初期化
    for p in [_IMG_LABEL_DIR, _IMG_PERIOD_DIR, _IMG_SUCCESS_DIR, _IMG_ERROR_DIR, _IMG_OLDLAYOUT_DIR ]:
        initialize_dir(p)

    #CSVのヘッダ
    header = ['period_img', 'max_height_img', 'filename', 'period', 'max_height', 'max_height_pixel']
    header.extend([str(i) for i in range(1, StepsImage.BIN_NUMBER + 1)])
    error_header = ['filename', 'exception_type', 'message', 'lineno']

    #Excelファイルの新規作成し、ヘッダを埋め込む
    if os.path.exists(_OUTPUT_EXCEL_PATH):
        os.remove(_OUTPUT_EXCEL_PATH)
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    for i, h in enumerate(header):
        c = ws.cell(row=1, column=i+1)
        c.value = h
    wb.save(_OUTPUT_EXCEL_PATH)

    #Excelを開いておく
    wb = openpyxl.load_workbook(_OUTPUT_EXCEL_PATH)
    ws = wb.worksheets[0]

    #棒グラフを読み取る
    results = []
    error_results_map = {}
    max_width_period, max_width_label = 0, 0
    write_xl_args = {}
    extract_image_details_args = []
    for index, file_ in enumerate(files):
        extract_image_details_args.append([index, file_, files, write_xl_args, error_results_map, error_header])

    for args in extract_image_details_args:
        extract_image_details(*args)
    # 同時実行(別にもっと早いでもない)
    # loop = asyncio.get_event_loop()
    # coroutines = [asyncio.to_thread(extract_image_details, *args) for args in extract_image_details_args]
    # loop.run_until_complete(asyncio.gather(*coroutines))

    for ni, (_, args) in enumerate(sorted(write_xl_args.items()), 1):
        max_width_label, max_width_period = write_xl_row(*args, ni, results, header, max_width_period, max_width_label, ws, wb)

    wb.save(_OUTPUT_EXCEL_PATH)

    #Excelのグラフ貼り付け部分のサイズ変更
    wb = openpyxl.load_workbook(_OUTPUT_EXCEL_PATH)
    ws = wb.worksheets[0]
    ws.column_dimensions['A'].width = max_width_period * 0.14
    ws.column_dimensions['B'].width = max_width_label * 0.14
    wb.save(_OUTPUT_EXCEL_PATH)

    #CSV出力
    encoding_ = 'cp932' #自分のPCがshift_jis使用であればcp932を使う必要がある（日本で多い）
    # encoding_ = 'utf8' #グローバルスタンダード
    write_csv(_RESULT_CSV_PATH, header, results, encoding_)
    error_results = [error_results_map[index] for index in sorted(error_results_map)]
    write_csv(_ERROR_CSV_PATH, error_header, error_results, encoding_)

    print(f'Elapsed time: {time.time() - start_time} second.')


if __name__== '__main__':
    main()
